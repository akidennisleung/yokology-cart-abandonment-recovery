#!/usr/bin/env python3
"""Generate formatted YOKOLOGY_Send_History_Log.xlsx from send_history.csv.

Usage:
    python generate_send_history_xlsx.py <send_history_csv> <output_xlsx>
"""
import sys
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill


def generate(csv_path: str, xlsx_path: str):
    history = pd.read_csv(csv_path, dtype=str)
    wb = Workbook()

    # Sheet 1: Full History
    ws1 = wb.active
    ws1.title = '發送記錄總表'

    headers = ['發送日期', '冷卻到期日', '全名', 'WhatsApp電話', '電郵', '意向程度',
               '客戶價值', 'GA4_事件數', '累積消費金額', '顧客ID', '發送批次', '備註']

    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)

    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    today = datetime.now().strftime('%Y-%m-%d')
    for row_idx, (_, record) in enumerate(history.iterrows(), 2):
        for col_idx, header in enumerate(headers, 1):
            val = record.get(header, '')
            cell = ws1.cell(row=row_idx, column=col_idx, value=val if pd.notna(val) else '')
            if header == '冷卻到期日' and pd.notna(val) and str(val) >= today:
                cell.font = Font(color='FF0000', bold=True)

    # Auto-width
    for col in range(1, len(headers) + 1):
        max_len = max(
            len(str(ws1.cell(row=r, column=col).value or ''))
            for r in range(1, len(history) + 2)
        )
        col_letter = ws1.cell(row=1, column=col).column_letter
        ws1.column_dimensions[col_letter].width = min(max_len + 4, 30)

    # Sheet 2: Cooldown Status
    ws2 = wb.create_sheet('冷卻狀態')
    ws2.cell(row=1, column=1, value='截至日期').font = Font(bold=True)
    ws2.cell(row=1, column=2, value=today)

    active = history[pd.to_datetime(history['冷卻到期日'], errors='coerce') >= pd.Timestamp(today)]
    ws2.cell(row=3, column=1, value='冷卻中電話號碼數').font = Font(bold=True)
    ws2.cell(row=3, column=2, value=len(active))

    # Breakdown by batch
    if '發送批次' in history.columns:
        for i, (batch, group) in enumerate(active.groupby('發送批次'), 4):
            ws2.cell(row=i, column=1, value=f'{batch}').font = Font(bold=True)
            ws2.cell(row=i, column=2, value=len(group))

    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 60

    wb.save(xlsx_path)
    print(f"[generate_send_history_xlsx] Saved: {xlsx_path}")


if __name__ == '__main__':
    if len(sys.argv) < 3:
        print(f"Usage: {sys.argv[0]} <send_history_csv> <output_xlsx>")
        sys.exit(1)
    generate(sys.argv[1], sys.argv[2])
