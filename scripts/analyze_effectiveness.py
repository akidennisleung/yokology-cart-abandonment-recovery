#!/usr/bin/env python3
"""Analyze cart abandonment recovery effectiveness by cross-referencing
send history with Shopline orders.

Usage:
    python analyze_effectiveness.py <send_history_csv> <orders_csv> <date_str> <output_dir>

Arguments:
    send_history_csv - Path to send_history.csv
    orders_csv       - Path to orders CSV (extracted from Shopline, see below)
    date_str         - Today's date in DDMon format (e.g. 15Apr)
    output_dir       - Directory for output report

Orders CSV format (one row per paid order):
    order_date,customer_name,email,amount
    2026-04-14,Rita Kwong,ritakwongsl@yahoo.com.hk,2272.36

The script matches send history records against orders placed AFTER the send
date (within 3-day attribution window) by email (primary) or name (fallback).
"""
import sys
import os
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


def load_orders(orders_csv):
    """Load and normalize orders CSV."""
    df = pd.read_csv(orders_csv, dtype=str)
    df['amount'] = pd.to_numeric(df['amount'], errors='coerce').fillna(0)
    df['email_lower'] = df['email'].str.lower().str.strip()
    df['name_lower'] = df['customer_name'].str.lower().str.strip()
    return df


def cross_reference(history, orders, attribution_days=3):
    """Match send records to subsequent orders within attribution window."""
    matches = []
    batches = history['發送批次'].unique()

    for batch in batches:
        batch_rows = history[history['發送批次'] == batch]
        if batch_rows.empty:
            continue

        send_date_str = batch_rows.iloc[0]['發送日期']
        try:
            send_date = pd.Timestamp(send_date_str)
        except Exception:
            continue

        window_end = send_date + timedelta(days=attribution_days)
        window_orders = orders[
            (pd.to_datetime(orders['order_date']) >= send_date) &
            (pd.to_datetime(orders['order_date']) <= window_end)
        ]

        for _, sent in batch_rows.iterrows():
            sent_email = str(sent.get('電郵', '')).lower().strip()
            sent_name = str(sent.get('全名', '')).lower().strip()

            # Email match (primary)
            email_hits = window_orders[window_orders['email_lower'] == sent_email]
            if len(email_hits) > 0 and sent_email:
                for _, order in email_hits.iterrows():
                    order_date = pd.Timestamp(order['order_date'])
                    matches.append({
                        'batch': batch,
                        'send_date': send_date_str,
                        'order_date': order['order_date'],
                        'days_to_convert': (order_date - send_date).days,
                        'customer_name': sent['全名'],
                        'email': sent.get('電郵', ''),
                        'phone': sent.get('WhatsApp電話', ''),
                        'tier': sent.get('意向程度', ''),
                        'customer_value': sent.get('客戶價值', ''),
                        'order_amount': order['amount'],
                        'match_method': 'email',
                    })
                continue

            # Name match (fallback)
            name_hits = window_orders[window_orders['name_lower'] == sent_name]
            if len(name_hits) > 0 and sent_name:
                for _, order in name_hits.iterrows():
                    order_date = pd.Timestamp(order['order_date'])
                    matches.append({
                        'batch': batch,
                        'send_date': send_date_str,
                        'order_date': order['order_date'],
                        'days_to_convert': (order_date - send_date).days,
                        'customer_name': sent['全名'],
                        'email': sent.get('電郵', ''),
                        'phone': sent.get('WhatsApp電話', ''),
                        'tier': sent.get('意向程度', ''),
                        'customer_value': sent.get('客戶價值', ''),
                        'order_amount': order['amount'],
                        'match_method': 'name',
                    })

    return pd.DataFrame(matches) if matches else pd.DataFrame()


def build_summary(history, matches_df):
    """Build per-batch and overall summary."""
    batches = history['發送批次'].unique()
    rows = []
    total_sent = 0
    total_converted = 0
    total_revenue = 0.0

    for batch in batches:
        batch_hist = history[history['發送批次'] == batch]
        sent = len(batch_hist)
        send_date = batch_hist.iloc[0]['發送日期']

        batch_matches = matches_df[matches_df['batch'] == batch] if len(matches_df) > 0 else pd.DataFrame()
        converted = len(batch_matches)
        revenue = batch_matches['order_amount'].sum() if len(batch_matches) > 0 else 0.0
        cvr = converted / sent * 100 if sent > 0 else 0

        total_sent += sent
        total_converted += converted
        total_revenue += revenue

        rows.append({
            'batch': batch, 'send_date': send_date, 'sent': sent,
            'converted': converted, 'cvr': cvr, 'revenue': revenue,
        })

    overall_cvr = total_converted / total_sent * 100 if total_sent > 0 else 0
    rows.append({
        'batch': '合計', 'send_date': '-', 'sent': total_sent,
        'converted': total_converted, 'cvr': overall_cvr, 'revenue': total_revenue,
    })
    return pd.DataFrame(rows), total_sent, total_converted, total_revenue, overall_cvr


def tier_analysis(matches_df):
    """Breakdown conversion by tier."""
    if matches_df.empty:
        return pd.DataFrame()
    grouped = matches_df.groupby('tier').agg(
        轉換人數=('customer_name', 'count'),
        總收益=('order_amount', 'sum'),
        平均訂單金額=('order_amount', 'mean'),
    ).reset_index()
    grouped.columns = ['意向程度', '轉換人數', '總收益 (HK$)', '平均訂單金額 (HK$)']
    return grouped


def value_analysis(matches_df):
    """Breakdown conversion by customer value."""
    if matches_df.empty:
        return pd.DataFrame()
    grouped = matches_df.groupby('customer_value').agg(
        轉換人數=('customer_name', 'count'),
        總收益=('order_amount', 'sum'),
    ).reset_index()
    grouped.columns = ['客戶價值', '轉換人數', '總收益 (HK$)']
    return grouped


def generate_report_xlsx(summary_df, matches_df, history, stats, date_str, output_dir):
    """Generate formatted XLSX effectiveness report."""
    total_sent, total_converted, total_revenue, overall_cvr = stats
    output_path = os.path.join(output_dir, f'YOKOLOGY_Effectiveness_Report_{date_str}.xlsx')

    wb = Workbook()
    hdr_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    hdr_font = Font(bold=True, color='FFFFFF', size=11)
    green = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    total_fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')

    # ── Sheet 1: Summary ──
    ws1 = wb.active
    ws1.title = '成效摘要'
    ws1['A1'] = f'YOKOLOGY 未結帳挽回流程 — 成效報告 ({date_str})'
    ws1['A1'].font = Font(bold=True, size=14, color='1F4E79')
    ws1.merge_cells('A1:G1')

    headers = ['廣播批次', '發送日期', '發送人數', '轉換人數', '轉換率', '帶來收益 (HK$)']
    for col, h in enumerate(headers, 1):
        c = ws1.cell(row=3, column=col, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal='center')

    for ri, (_, row) in enumerate(summary_df.iterrows(), 4):
        is_total = row['batch'] == '合計'
        fill = total_fill if is_total else green
        vals = [row['batch'], row['send_date'], row['sent'],
                row['converted'], f"{row['cvr']:.1f}%", f"HK${row['revenue']:,.2f}"]
        for ci, v in enumerate(vals, 1):
            c = ws1.cell(row=ri, column=ci, value=v)
            c.fill = fill
            c.alignment = Alignment(horizontal='center')
            if is_total:
                c.font = Font(bold=True)

    # KPIs
    kpi_row = len(summary_df) + 5
    ws1.cell(row=kpi_row, column=1, value='關鍵指標').font = Font(bold=True, size=12, color='1F4E79')
    kpis = [
        ('總發送人數', f'{total_sent} 人'),
        ('總轉換人數', f'{total_converted} 人'),
        ('整體轉換率', f'{overall_cvr:.1f}%'),
        ('總帶來收益', f'HK${total_revenue:,.2f}'),
        ('平均每位轉換客戶收益', f'HK${total_revenue/total_converted:,.2f}' if total_converted > 0 else 'N/A'),
    ]
    for i, (label, val) in enumerate(kpis, kpi_row + 1):
        ws1.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws1.cell(row=i, column=2, value=val).font = Font(color='1F4E79', bold=True, size=11)

    # Tier breakdown
    tier_row = kpi_row + len(kpis) + 2
    tier_df = tier_analysis(matches_df)
    if not tier_df.empty:
        ws1.cell(row=tier_row, column=1, value='按意向程度分析').font = Font(bold=True, size=12, color='1F4E79')
        for ci, h in enumerate(tier_df.columns, 1):
            c = ws1.cell(row=tier_row + 1, column=ci, value=h)
            c.font = hdr_font
            c.fill = hdr_fill
        for ri, (_, row) in enumerate(tier_df.iterrows(), tier_row + 2):
            for ci, col in enumerate(tier_df.columns, 1):
                v = row[col]
                if isinstance(v, float):
                    v = f'HK${v:,.2f}'
                ws1.cell(row=ri, column=ci, value=v)

    # Value breakdown
    val_row = tier_row + len(tier_df) + 4
    val_df = value_analysis(matches_df)
    if not val_df.empty:
        ws1.cell(row=val_row, column=1, value='按客戶價值分析').font = Font(bold=True, size=12, color='1F4E79')
        for ci, h in enumerate(val_df.columns, 1):
            c = ws1.cell(row=val_row + 1, column=ci, value=h)
            c.font = hdr_font
            c.fill = hdr_fill
        for ri, (_, row) in enumerate(val_df.iterrows(), val_row + 2):
            for ci, col in enumerate(val_df.columns, 1):
                v = row[col]
                if isinstance(v, float):
                    v = f'HK${v:,.2f}'
                ws1.cell(row=ri, column=ci, value=v)

    for col in range(1, 8):
        ws1.column_dimensions[get_column_letter(col)].width = 22

    # ── Sheet 2: Conversion Details ──
    ws2 = wb.create_sheet('轉換明細')
    detail_headers = ['廣播批次', '發送日期', '下單日期', '轉換天數', '顧客姓名',
                      '電郵', '電話', '意向程度', '客戶價值', '訂單金額 (HK$)']
    for ci, h in enumerate(detail_headers, 1):
        c = ws2.cell(row=1, column=ci, value=h)
        c.font = hdr_font
        c.fill = hdr_fill

    if not matches_df.empty:
        for ri, (_, row) in enumerate(matches_df.iterrows(), 2):
            vals = [row['batch'], row['send_date'], row['order_date'],
                    f"+{row['days_to_convert']}天" if row['days_to_convert'] > 0 else '當天',
                    row['customer_name'], row['email'], row['phone'],
                    row['tier'], row['customer_value'], f"HK${row['order_amount']:,.2f}"]
            for ci, v in enumerate(vals, 1):
                ws2.cell(row=ri, column=ci, value=v)
    else:
        ws2.cell(row=2, column=1, value='暫無轉換記錄')

    for col in range(1, 11):
        ws2.column_dimensions[get_column_letter(col)].width = 18

    # ── Sheet 3: Non-converters ──
    ws3 = wb.create_sheet('未轉換名單')
    nc_headers = ['廣播批次', '發送日期', '冷卻到期日', '顧客姓名', '電郵', '電話', '意向程度', '客戶價值']
    for ci, h in enumerate(nc_headers, 1):
        c = ws3.cell(row=1, column=ci, value=h)
        c.font = hdr_font
        c.fill = hdr_fill

    converted_emails = set(matches_df['email'].str.lower().str.strip()) if len(matches_df) > 0 else set()
    non_conv = history[~history['電郵'].str.lower().str.strip().isin(converted_emails)]
    for ri, (_, row) in enumerate(non_conv.iterrows(), 2):
        vals = [row.get('發送批次',''), row.get('發送日期',''), row.get('冷卻到期日',''),
                row.get('全名',''), row.get('電郵',''), row.get('WhatsApp電話',''),
                row.get('意向程度',''), row.get('客戶價值','')]
        for ci, v in enumerate(vals, 1):
            ws3.cell(row=ri, column=ci, value=v)

    for col in range(1, 9):
        ws3.column_dimensions[get_column_letter(col)].width = 22

    wb.save(output_path)
    return output_path


def run(send_history_csv, orders_csv, date_str, output_dir):
    """Main entry point."""
    os.makedirs(output_dir, exist_ok=True)

    # Load data
    history = pd.read_csv(send_history_csv, dtype=str)
    # Deduplicate by phone+batch
    history = history.drop_duplicates(subset=['WhatsApp電話', '發送批次'], keep='first')
    orders = load_orders(orders_csv)

    # Cross-reference
    matches = cross_reference(history, orders, attribution_days=3)

    # Build summary
    summary_df, total_sent, total_converted, total_revenue, overall_cvr = build_summary(history, matches)

    # Print console summary
    print(f"\n{'='*60}")
    print(f"成效分析 — {date_str}")
    print(f"{'='*60}")
    for _, row in summary_df.iterrows():
        print(f"  {row['batch']:<16} | 發送 {row['sent']:>3} | 轉換 {row['converted']:>2} | CVR {row['cvr']:>5.1f}% | HK${row['revenue']:>10,.2f}")

    if not matches.empty:
        print(f"\n轉換明細:")
        for _, m in matches.iterrows():
            days_label = f"+{m['days_to_convert']}天" if m['days_to_convert'] > 0 else '當天'
            print(f"  {m['customer_name']:<24} | {m['batch']} → {m['order_date']} ({days_label}) | HK${m['order_amount']:,.2f} | {m['tier']}")

    # Generate report
    stats = (total_sent, total_converted, total_revenue, overall_cvr)
    report_path = generate_report_xlsx(summary_df, matches, history, stats, date_str, output_dir)
    print(f"\n報告已生成: {report_path}")

    return summary_df, matches, report_path


if __name__ == '__main__':
    if len(sys.argv) < 5:
        print(f"Usage: {sys.argv[0]} <send_history_csv> <orders_csv> <date_str> <output_dir>")
        sys.exit(1)
    run(sys.argv[1], sys.argv[2], sys.argv[3], sys.argv[4])
