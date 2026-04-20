#!/usr/bin/env python3
"""Classify intent tiers, match with Shopline, apply cooldown, and generate output files.

Usage:
    python classify_and_match.py <ga4_parquet> <shopline_xls> <send_history_csv> <date_str> <output_dir>

Arguments:
    ga4_csv         - Parsed GA4 data CSV (output of parse_ga4_csv.py)
    shopline_xls    - Shopline customer report (.xls)
    send_history_csv - Send history CSV (will be created if missing)
    date_str        - Today's date in format DDMon (e.g. 14Apr)
    output_dir      - Directory for output files
"""
import sys
import os
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill


def classify_tier(row):
    """Classify non-purchaser into Tier A/B/C based on V2 criteria."""
    ec = row['event_count']
    ke = row.get('key_events', 0)
    if ec >= 40 and ke > 0:
        return 'Tier A'
    elif ec >= 20:
        return 'Tier B'
    elif ec >= 5:
        return 'Tier C'
    return None


def classify_value(spent_str):
    """Classify customer value based on cumulative spending."""
    try:
        spent = float(str(spent_str).replace('HK$', '').replace(',', '').replace('$', '').strip())
    except (ValueError, TypeError):
        spent = 0
    if spent >= 100000:
        return 'VIP客戶', spent
    elif spent >= 50000:
        return '重要客戶', spent
    elif spent >= 10000:
        return '活躍客戶', spent
    elif spent > 0:
        return '一般客戶', spent
    return '潛在新客', spent


def suggest_message(value):
    """Suggest WhatsApp message type based on customer value."""
    mapping = {
        'VIP客戶': 'VIP關懷問候 + 專屬優惠',
        '重要客戶': 'VIP關懷問候 + 專屬優惠',
        '活躍客戶': '新品推薦 + 回購優惠',
        '一般客戶': '產品推薦 + 滿額免運',
        '潛在新客': '品牌介紹 + 首購優惠碼',
    }
    return mapping.get(value, '品牌介紹 + 首購優惠碼')


def get_best_phone(row):
    """Extract best phone number from Shopline columns (priority order)."""
    for col in ['聯絡電話', '會員綁定手機號碼', '收件人電話']:
        val = row.get(col)
        if pd.notna(val) and str(val).strip() not in ('', 'nan', 'None'):
            return str(val).strip()
    return ''


def clean_phone(phone_str):
    """Clean phone number: remove non-digits, keep last 8 digits for HK numbers."""
    import re
    digits = re.sub(r'[^\d]', '', str(phone_str))
    if len(digits) > 8:
        digits = digits[-8:]
    return digits if len(digits) == 8 else ''


def run(ga4_csv_path, shopline_xls, send_history_csv, date_str, output_dir):
    """Main workflow: classify, match, dedup, output."""
    os.makedirs(output_dir, exist_ok=True)
    today = datetime.now().strftime('%Y-%m-%d')
    cooldown_expiry = (datetime.now() + timedelta(days=10)).strftime('%Y-%m-%d')
    summary = {}

    # --- Step 1: Load GA4 data ---
    ga4 = pd.read_csv(ga4_csv_path)
    for col in ['event_count', 'sessions', 'avg_session_duration', 'purchase_revenue', 'transactions', 'active_users', 'key_events']:
        if col in ga4.columns:
            ga4[col] = pd.to_numeric(ga4[col], errors='coerce').fillna(0)
    summary['ga4_total'] = len(ga4)

    # --- Step 2: Filter non-purchasers and classify tiers ---
    non_purchasers = ga4[
        (ga4['purchase_revenue'] == 0) & (ga4['transactions'] == 0)
    ].copy()
    non_purchasers['tier'] = non_purchasers.apply(classify_tier, axis=1)
    tiered = non_purchasers[non_purchasers['tier'].notna()].copy()

    summary['tier_a'] = len(tiered[tiered['tier'] == 'Tier A'])
    summary['tier_b'] = len(tiered[tiered['tier'] == 'Tier B'])
    summary['tier_c'] = len(tiered[tiered['tier'] == 'Tier C'])

    # --- Step 3: Match with Shopline ---
    shopline = pd.read_excel(shopline_xls, dtype=str)
    shopline['clean_id'] = shopline['顧客 ID'].astype(str).str.strip()

    matched = tiered.merge(shopline, left_on='user_id', right_on='clean_id', how='inner')

    # Identify unmatched IDs (new member alert)
    matched_ids = set(matched['user_id'])
    all_tiered_ids = set(tiered['user_id'])
    unmatched_ids = all_tiered_ids - matched_ids
    summary['unmatched_ids'] = list(unmatched_ids)

    # Extract best phone and filter
    matched['best_phone'] = matched.apply(get_best_phone, axis=1)
    matched['clean_phone'] = matched['best_phone'].apply(clean_phone)
    matched = matched[matched['clean_phone'] != ''].copy()

    # Classify customer value
    value_results = matched['累積金額'].apply(classify_value)
    matched['客戶價值'] = [v[0] for v in value_results]
    matched['消費金額數值'] = [v[1] for v in value_results]
    matched['建議訊息類型'] = matched['客戶價值'].apply(suggest_message)

    # Sort by tier priority then event count
    tier_order = {'Tier A': 0, 'Tier B': 1, 'Tier C': 2}
    matched['tier_sort'] = matched['tier'].map(tier_order)
    matched = matched.sort_values(['tier_sort', 'event_count'], ascending=[True, False])

    summary['matched_with_phone'] = len(matched)

    # --- Step 4: 10-day cooldown dedup ---
    excluded_cooldown = 0
    if os.path.exists(send_history_csv):
        history = pd.read_csv(send_history_csv, dtype=str)
        history['冷卻到期日'] = pd.to_datetime(history['冷卻到期日'], errors='coerce')
        active_cooldown = history[history['冷卻到期日'] >= pd.Timestamp(today)]
        cooldown_phones = set(active_cooldown['WhatsApp電話'].astype(str).str.strip())
        before = len(matched)
        matched = matched[~matched['clean_phone'].isin(cooldown_phones)]
        excluded_cooldown = before - len(matched)

    summary['excluded_cooldown'] = excluded_cooldown

    # --- Step 4b: Unsubscribe exclusion ---
    excluded_unsub = 0
    unsub_path = os.path.join(os.path.dirname(send_history_csv), 'unsubscribed.xlsx')
    if os.path.exists(unsub_path):
        unsub_df = pd.read_excel(unsub_path, dtype=str, header=None)
        # The file has two columns: country code (852) and phone number
        # First row may be a header like '取消訂閱名單', skip non-numeric rows
        if len(unsub_df.columns) >= 2:
            phone_col = unsub_df.iloc[:, 1].dropna().astype(str).str.strip()
        else:
            phone_col = unsub_df.iloc[:, 0].dropna().astype(str).str.strip()
        unsub_phones = set(p for p in phone_col if p.isdigit() and len(p) == 8)
        before = len(matched)
        matched = matched[~matched['clean_phone'].isin(unsub_phones)]
        excluded_unsub = before - len(matched)
    summary['excluded_unsub'] = excluded_unsub

    # --- Step 5: Remove duplicates by phone ---
    matched = matched.drop_duplicates(subset='clean_phone', keep='first')

    # --- Step 6: Build output ---
    output = matched[['全名', 'clean_phone', '電郵', 'tier', '客戶價值',
                       'event_count', 'key_events', 'sessions',
                       '累積金額', '訂單數', '建議訊息類型', 'user_id']].copy()
    output.columns = ['全名', 'WhatsApp電話', '電郵', '意向程度', '客戶價值',
                       'GA4_事件數', 'GA4_關鍵事件數', 'GA4_工作階段',
                       '累積消費金額', '總訂單數', '建議訊息類型', '顧客ID']

    summary['final_count'] = len(output)

    # --- File 1: Internal Review XLSX ---
    internal_path = os.path.join(output_dir, f'YOKOLOGY_WhatsApp_Broadcast_{date_str}.xlsx')
    output.to_excel(internal_path, index=False)

    # --- File 2: WhatsApp Upload XLSX ---
    upload_path = os.path.join(output_dir, f'YOKOLOGY_WhatsApp_Upload_{date_str}.xlsx')
    wb = Workbook()
    ws = wb.active
    ws.title = 'WhatsApp Upload'
    ws.merge_cells('A1:B1')
    ws['A1'] = 'WhatsApp 聯絡人手機號碼上傳說明： \n\n請於下方欄位中填入需要上傳的聯絡人手機號碼資訊，其中手機號國碼和手機號碼均為必填項目，若聯絡人的手機號碼資訊未完整填寫將上傳失敗'
    ws['A1'].font = Font(bold=True, size=10)
    ws['A1'].alignment = Alignment(wrap_text=True)
    ws['A2'] = '手機號國碼\n範例：+852'
    ws['A2'].font = Font(bold=True)
    ws['A2'].alignment = Alignment(wrap_text=True)
    ws['B2'] = '手機號碼'
    ws['B2'].font = Font(bold=True)
    for i, (_, row) in enumerate(output.iterrows(), start=3):
        ws[f'A{i}'] = 852
        ws[f'B{i}'] = int(row['WhatsApp電話']) if str(row['WhatsApp電話']).isdigit() else row['WhatsApp電話']
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    wb.save(upload_path)

    # --- Step 7: Update send history ---
    new_records = []
    for _, row in output.iterrows():
        new_records.append({
            '發送日期': today,
            '冷卻到期日': cooldown_expiry,
            '全名': row['全名'],
            'WhatsApp電話': str(row['WhatsApp電話']),
            '電郵': row['電郵'],
            '意向程度': row['意向程度'],
            '客戶價值': row['客戶價值'],
            'GA4_事件數': row['GA4_事件數'],
            '累積消費金額': row['累積消費金額'],
            '顧客ID': row['顧客ID'],
            '發送批次': f'{date_str}_Batch1',
            '備註': ''
        })

    new_df = pd.DataFrame(new_records)
    if os.path.exists(send_history_csv):
        existing = pd.read_csv(send_history_csv, dtype=str)
        combined = pd.concat([existing, new_df], ignore_index=True)
    else:
        combined = new_df
    combined.to_csv(send_history_csv, index=False)

    # --- Print summary ---
    print(f"\n{'='*60}")
    print(f"執行摘要 — {date_str}")
    print(f"{'='*60}")
    print(f"GA4 有效用戶:           {summary['ga4_total']}")
    print(f"高意向未結帳 (Tier A):  {summary['tier_a']}")
    print(f"高意向未結帳 (Tier B):  {summary['tier_b']}")
    print(f"一般意向未結帳 (Tier C): {summary['tier_c']}")
    print(f"Shopline 匹配 (有電話): {summary['matched_with_phone']}")
    print(f"排除：10天冷卻期內:     -{summary['excluded_cooldown']}")
    print(f"排除：取消訂閱:         -{summary['excluded_unsub']}")
    print(f"最終廣播名單:           {summary['final_count']}")
    if summary['unmatched_ids']:
        print(f"\n⚠️ 新註冊會員提醒 ({len(summary['unmatched_ids'])} 位):")
        for uid in summary['unmatched_ids']:
            print(f"  {uid}")

    print(f"\n輸出檔案:")
    print(f"  內部查閱版: {internal_path}")
    print(f"  直接上傳版: {upload_path}")
    print(f"  發送記錄:   {send_history_csv}")

    return summary, output


if __name__ == '__main__':
    if len(sys.argv) < 6:
        print(f"Usage: {sys.argv[0]} <ga4_csv> <shopline_xls> <send_history_csv> <date_str> <output_dir>")
        sys.exit(1)
    run(sys.argv[1], sys.argv[2], sys.argv[3], sys.argv[4], sys.argv[5])
